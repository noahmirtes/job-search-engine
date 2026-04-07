"""Report export flow: query scored jobs, write workbook tabs, track exports."""

from __future__ import annotations

import json
import sqlite3
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import WorkerConfig
from app.db import utc_now_iso
from app.worker_logging import get_logger


PASTEL_SECTION_STYLES = {
    "metadata": {
        "header_fill": PatternFill(fill_type="solid", fgColor="D9EAF7"),
        "body_fill": PatternFill(fill_type="solid", fgColor="F5FAFE"),
    },
    "details": {
        "header_fill": PatternFill(fill_type="solid", fgColor="E3F1DD"),
        "body_fill": PatternFill(fill_type="solid", fgColor="F7FBF4"),
    },
    "apply": {
        "header_fill": PatternFill(fill_type="solid", fgColor="F8E199"),
        "body_fill": PatternFill(fill_type="solid", fgColor="FFFECC"),
    },
}
HEADER_FONT = Font(bold=True, color="3F4E5A")
BODY_FONT = Font(color="4A4A4A")
THIN_BORDER = Border(
    left=Side(style="thin", color="E4E8EC"),
    right=Side(style="thin", color="E4E8EC"),
    top=Side(style="thin", color="E4E8EC"),
    bottom=Side(style="thin", color="E4E8EC"),
)
DEFAULT_ALIGNMENT = Alignment(vertical="top")
HEADER_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=False)
WRAPPED_COLUMNS = {"Description", "Qualifications", "Source Queries"}
COLUMN_WIDTH_CAPS = {
    "Job ID": 10,
    "Score": 10,
    "Fit Recommendation": 22,
    "Source Queries": 30,
    "Title": 36,
    "Company": 28,
    "Location": 24,
    "Description": 60,
    "Date Posted": 14,
    "Schedule Type": 18,
    "Qualifications": 44,
    "Extensions": 26,
    "Detected Extensions": 28,
}
LOGGER = get_logger("reporting")


@dataclass(frozen=True)
class ReportRunSummary:
    """Run summary returned after one report export."""

    export_id: int
    export_path: Path
    new_count: int
    all_count: int
    all_jobs_list_count: int


@dataclass(frozen=True)
class ReportSheetPayload:
    """Sheet-ready table plus hyperlink metadata for apply columns."""

    dataframe: pd.DataFrame
    apply_links_by_row: list[dict[str, str]]


def generate_report(
    connection: sqlite3.Connection,
    config: WorkerConfig,
) -> ReportRunSummary:
    """Generate one Excel report and record export metadata in the DB."""
    scoring_version = config.scoring_config.version
    threshold = config.scoring_config.report.threshold
    include_all_jobs_list = config.scoring_config.report.include_all_jobs_list
    LOGGER.info(
        "Report generation start: version=%s threshold=%s include_all_jobs_list=%s",
        scoring_version,
        threshold,
        include_all_jobs_list,
    )

    all_rows = _fetch_scored_rows(
        connection=connection,
        scoring_version=scoring_version,
        threshold=threshold,
    )

    last_exported_at = _get_latest_exported_at(connection)
    if last_exported_at is None:
        new_rows = all_rows
    else:
        new_rows = [
            row for row in all_rows
            if _as_text(row["first_seen_at"]) and _as_text(row["first_seen_at"]) > last_exported_at
        ]

    all_jobs_list_rows: list[sqlite3.Row] = []
    if include_all_jobs_list:
        all_jobs_list_rows = _fetch_all_jobs_rows(
            connection=connection,
            scoring_version=scoring_version,
        )

    report_file_name = _build_report_file_name()
    report_path = config.paths.report_export_dir / report_file_name
    _write_report_workbook(
        report_path=report_path,
        new_rows=new_rows,
        all_rows=all_rows,
        all_jobs_list_rows=all_jobs_list_rows,
        include_all_jobs_list=include_all_jobs_list,
    )

    export_id = _insert_export(connection, report_file_name)
    _insert_export_jobs(
        connection=connection,
        export_id=export_id,
        job_ids=[int(row["job_id"]) for row in new_rows],
    )

    summary = ReportRunSummary(
        export_id=export_id,
        export_path=report_path,
        new_count=len(new_rows),
        all_count=len(all_rows),
        all_jobs_list_count=len(all_jobs_list_rows),
    )
    LOGGER.info(
        "Report generation complete: export_id=%s path=%s new_rows=%s all_rows=%s all_jobs_list_rows=%s",
        summary.export_id,
        summary.export_path,
        summary.new_count,
        summary.all_count,
        summary.all_jobs_list_count,
    )
    return summary


def _fetch_scored_rows(
    *,
    connection: sqlite3.Connection,
    scoring_version: str,
    threshold: float,
) -> list[sqlite3.Row]:
    """Load threshold-qualified scored jobs for the configured scoring version."""
    rows = connection.execute(
        """
        SELECT
            jobs.id AS job_id,
            jobs.title,
            jobs.company,
            jobs.location,
            jobs.description,
            jobs.date_posted,
            jobs.schedule_type,
            jobs.qualifications_text,
            jobs.extensions_json,
            jobs.detected_extensions_json,
            jobs.apply_options_json,
            jobs.query_names_json,
            jobs.first_seen_at,
            jobs.last_seen_at,
            job_scores.total_score,
            job_scores.fit_recommendation,
            job_scores.scored_at
        FROM jobs
        JOIN job_scores
            ON job_scores.job_id = jobs.id
        WHERE job_scores.scoring_version = ?
          AND job_scores.scoring_status = 'ok'
          AND job_scores.total_score >= ?
          AND jobs.is_scorable = 1
        ORDER BY job_scores.total_score DESC, jobs.last_seen_at DESC, jobs.id DESC
        """,
        (scoring_version, threshold),
    ).fetchall()
    return [row for row in rows if isinstance(row, sqlite3.Row)]


def _fetch_all_jobs_rows(
    *,
    connection: sqlite3.Connection,
    scoring_version: str,
) -> list[sqlite3.Row]:
    """Load all jobs, attaching score when available for the configured version."""
    rows = connection.execute(
        """
        SELECT
            jobs.id AS job_id,
            jobs.title,
            jobs.company,
            jobs.location,
            jobs.description,
            jobs.date_posted,
            jobs.schedule_type,
            jobs.qualifications_text,
            jobs.extensions_json,
            jobs.detected_extensions_json,
            jobs.apply_options_json,
            jobs.query_names_json,
            jobs.first_seen_at,
            jobs.last_seen_at,
            job_scores.total_score,
            job_scores.fit_recommendation,
            job_scores.scored_at
        FROM jobs
        LEFT JOIN job_scores
            ON job_scores.job_id = jobs.id
           AND job_scores.scoring_version = ?
           AND job_scores.scoring_status = 'ok'
        WHERE jobs.is_scorable = 1
          AND NOT EXISTS (
              SELECT 1
              FROM job_scores AS blacklisted_scores
              WHERE blacklisted_scores.job_id = jobs.id
                AND blacklisted_scores.scoring_version = ?
                AND blacklisted_scores.scoring_status = 'blacklisted'
          )
        ORDER BY
            CASE WHEN job_scores.total_score IS NULL THEN 1 ELSE 0 END ASC,
            job_scores.total_score DESC,
            jobs.last_seen_at DESC,
            jobs.id DESC
        """,
        (scoring_version, scoring_version),
    ).fetchall()
    return [row for row in rows if isinstance(row, sqlite3.Row)]


def _get_latest_exported_at(connection: sqlite3.Connection) -> str | None:
    """Return the timestamp of the most recent export, if any."""
    row = connection.execute(
        "SELECT exported_at FROM exports ORDER BY exported_at DESC, id DESC LIMIT 1"
    ).fetchone()
    if row is None:
        return None
    return _as_text(row["exported_at"] if isinstance(row, sqlite3.Row) else row[0])


def _write_report_workbook(
    *,
    report_path: Path,
    new_rows: list[sqlite3.Row],
    all_rows: list[sqlite3.Row],
    all_jobs_list_rows: list[sqlite3.Row],
    include_all_jobs_list: bool,
) -> None:
    """Write Excel workbook with required tabs and dynamic apply columns."""
    report_path.parent.mkdir(parents=True, exist_ok=True)
    sheet_payloads: list[tuple[str, ReportSheetPayload]] = [
        ("new", _build_sheet_payload(new_rows)),
        ("all", _build_sheet_payload(all_rows)),
    ]
    if include_all_jobs_list:
        sheet_payloads.append(("all_jobs_list", _build_sheet_payload(all_jobs_list_rows)))

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        for sheet_name, payload in sheet_payloads:
            payload.dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            _style_worksheet(worksheet)
            _apply_sheet_hyperlinks(worksheet, payload)


def _build_sheet_payload(rows: list[sqlite3.Row]) -> ReportSheetPayload:
    """Build report rows plus hyperlink metadata for one workbook sheet."""
    max_apply_options = 0
    parsed_options_by_job: dict[int, list[tuple[str, str]]] = {}

    for row in rows:
        job_id = int(row["job_id"])
        options = _parse_apply_options(row["apply_options_json"])
        parsed_options_by_job[job_id] = options
        max_apply_options = max(max_apply_options, len(options))

    columns = [
        "Job ID",
        "Score",
        "Fit Recommendation",
        "Source Queries",
        "Title",
        "Company",
        "Location",
        "Description",
        "Date Posted",
        "Schedule Type",
        "Qualifications",
        "Extensions",
        "Detected Extensions",
    ]
    for index in range(1, max_apply_options + 1):
        columns.append(f"Apply Location {index}")

    records: list[dict[str, Any]] = []
    apply_links_by_row: list[dict[str, str]] = []
    for row in rows:
        job_id = int(row["job_id"])
        options = parsed_options_by_job[job_id]
        row_links: dict[str, str] = {}

        record: dict[str, Any] = {
            "Job ID": job_id,
            "Score": row["total_score"],
            "Fit Recommendation": _as_text(row["fit_recommendation"]) or "",
            "Source Queries": _format_query_names(row["query_names_json"]),
            "Title": _as_text(row["title"]) or "",
            "Company": _as_text(row["company"]) or "",
            "Location": _as_text(row["location"]) or "",
            "Description": _as_text(row["description"]) or "",
            "Date Posted": _as_text(row["date_posted"]) or "",
            "Schedule Type": _as_text(row["schedule_type"]) or "",
            "Qualifications": _as_text(row["qualifications_text"]) or "",
            "Extensions": _as_text(row["extensions_json"]) or "",
            "Detected Extensions": _as_text(row["detected_extensions_json"]) or "",
        }

        for index in range(max_apply_options):
            column_index = index + 1
            if index < len(options):
                apply_location, apply_link = options[index]
            else:
                apply_location, apply_link = "", ""
            column_name = f"Apply Location {column_index}"
            display_value = apply_location or (f"Apply Link {column_index}" if apply_link else "")
            record[column_name] = display_value
            if apply_link:
                row_links[column_name] = apply_link

        records.append(record)
        apply_links_by_row.append(row_links)

    return ReportSheetPayload(
        dataframe=pd.DataFrame(records, columns=columns),
        apply_links_by_row=apply_links_by_row,
    )


def _apply_sheet_hyperlinks(worksheet: Any, payload: ReportSheetPayload) -> None:
    """Attach hyperlink targets to apply-location cells after sheet write."""
    header_names = [cell.value for cell in worksheet[1]]
    header_index = {
        str(name): position
        for position, name in enumerate(header_names, start=1)
        if isinstance(name, str)
    }

    for row_offset, row_links in enumerate(payload.apply_links_by_row, start=2):
        for column_name, link in row_links.items():
            column_index = header_index.get(column_name)
            if column_index is None:
                continue
            cell = worksheet.cell(row=row_offset, column=column_index)
            cell.hyperlink = link
            cell.font = Font(color="0563C1", underline="single")


def _style_worksheet(worksheet: Any) -> None:
    """Apply pastel section styling and usability formatting to one sheet."""
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 24

    header_names = [cell.value for cell in worksheet[1]]
    for column_index, header_value in enumerate(header_names, start=1):
        header = header_value if isinstance(header_value, str) else ""
        family = _column_family(header)
        fills = PASTEL_SECTION_STYLES[family]
        column_letter = get_column_letter(column_index)

        header_cell = worksheet.cell(row=1, column=column_index)
        header_cell.fill = fills["header_fill"]
        header_cell.font = HEADER_FONT
        header_cell.alignment = HEADER_ALIGNMENT
        header_cell.border = THIN_BORDER

        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.fill = fills["body_fill"]
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGNMENT if header in WRAPPED_COLUMNS else DEFAULT_ALIGNMENT
            cell.border = THIN_BORDER

        worksheet.column_dimensions[column_letter].width = _compute_column_width(
            worksheet,
            column_index,
            header,
        )


def _column_family(header: str) -> str:
    """Group columns into styling families."""
    if header.startswith("Apply Location "):
        return "apply"
    if header in {
        "Job ID",
        "Score",
        "Fit Recommendation",
        "Source Queries",
        "Date Posted",
        "Schedule Type",
    }:
        return "metadata"
    return "details"


def _compute_column_width(
    worksheet: Any,
    column_index: int,
    header: str,
) -> float:
    """Auto-size one column with a per-column max width cap."""
    width_cap = 20 if header.startswith("Apply Location ") else COLUMN_WIDTH_CAPS.get(header, 24)
    max_length = len(header)

    for row_index in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_index, column=column_index).value
        if value is None:
            continue
        max_length = max(max_length, len(str(value).strip()))

    return min(max(max_length + 2, 10), width_cap)


def _parse_apply_options(raw_apply_options_json: str | None) -> list[tuple[str, str]]:
    """Extract ordered (location, link) tuples from apply_options_json."""
    if not raw_apply_options_json:
        return []
    try:
        payload = json.loads(raw_apply_options_json)
    except json.JSONDecodeError:
        return []

    if not isinstance(payload, list):
        return []

    items: list[tuple[str, str]] = []
    for option in payload:
        if not isinstance(option, dict):
            continue

        location = (
            _as_text(option.get("title"))
            or _as_text(option.get("via"))
            or _as_text(option.get("source"))
            or ""
        )
        link = (
            _as_text(option.get("link"))
            or _as_text(option.get("apply_link"))
            or _as_text(option.get("url"))
            or ""
        )

        if location or link:
            items.append((location, link))

    return items


def _format_query_names(raw_query_names_json: str | None) -> str:
    """Render stored query names JSON as a report-friendly string."""
    if not raw_query_names_json:
        return ""
    try:
        payload = json.loads(raw_query_names_json)
    except json.JSONDecodeError:
        return ""

    if not isinstance(payload, list):
        return ""

    query_names = []
    for value in payload:
        text = _as_text(value)
        if text:
            query_names.append(text)

    return ", ".join(query_names)


def _build_report_file_name() -> str:
    """Build stable timestamped report filename."""
    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    return f"job_report_{stamp}.xlsx"


def _insert_export(connection: sqlite3.Connection, report_file_name: str) -> int:
    """Insert an exports row and return new export id."""
    cursor = connection.execute(
        "INSERT INTO exports (exported_at, export_file_name) VALUES (?, ?)",
        (utc_now_iso(), report_file_name),
    )
    return int(cursor.lastrowid)


def _insert_export_jobs(
    *,
    connection: sqlite3.Connection,
    export_id: int,
    job_ids: list[int],
) -> None:
    """Link jobs surfaced in the new tab to the export run."""
    if not job_ids:
        return
    connection.executemany(
        "INSERT OR IGNORE INTO export_jobs (export_id, job_id) VALUES (?, ?)",
        [(export_id, job_id) for job_id in job_ids],
    )


def _as_text(value: Any) -> str | None:
    """Normalize optional values to trimmed text."""
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return None
